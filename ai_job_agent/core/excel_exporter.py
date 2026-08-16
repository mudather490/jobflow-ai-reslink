import os
import csv
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

from config import OUTPUT_DIR
from core.global_employers import get_all_global_employers

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def sanitize_excel_cell(val: Any) -> str:
    """
    Prevents CSV / Formula Injection (CWE-1236).
    Prepends a single quote if string starts with formula characters (=, +, -, @, tab, cr).
    """
    if val is None:
        return ""
    raw = str(val)
    if not raw:
        return ""
    if raw[0] in ("=", "+", "-", "@", "\t", "\r") or raw.startswith("%09") or raw.startswith("%0D"):
        return f"'{raw.strip()}"
    s = raw.strip()
    if s and s[0] in ("=", "+", "-", "@"):
        return f"'{s}"
    return s


class CompanyIntelligenceExcelExporter:
    """
    Generates clean, professional Excel (.xlsx) and CSV spreadsheets
    containing applied company profiles, role requirements, and key intelligence
    to help candidates prepare for interviews.
    """

    @classmethod
    def get_company_overview(cls, company_name: str, job_title: str, description: str = "") -> str:
        """
        Synthesizes a clean company overview from global employers database or job description.
        """
        clean_name = company_name.strip()
        global_map = get_all_global_employers()

        for c_name, c_loc, c_notes in global_map:
            if c_name.lower() in clean_name.lower() or clean_name.lower() in c_name.lower():
                return f"{c_loc} • {c_notes}"

        # Extract summary from job description if available
        if description:
            desc_lines = [l.strip() for l in description.split("\n") if len(l.strip()) > 20]
            if desc_lines:
                return desc_lines[0][:120] + "..." if len(desc_lines[0]) > 120 else desc_lines[0]

        return f"Technology & Business Operations • Hiring for {job_title}"

    @classmethod
    def export_to_excel(
        cls,
        applications: List[Dict[str, Any]],
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Builds a styled Excel spreadsheet (.xlsx) with clean company intelligence.
        """
        out_file = output_path or (OUTPUT_DIR / "Company_Applications_Tracker.xlsx")
        out_file.parent.mkdir(parents=True, exist_ok=True)

        if not OPENPYXL_AVAILABLE:
            # Fallback to CSV if openpyxl is not present
            csv_path = out_file.with_suffix(".csv")
            cls.export_to_csv(applications, csv_path)
            return csv_path

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Company Intelligence Tracker"

        # Headers
        headers = [
            "Company Name",
            "Job Title / Role",
            "Company Overview & Domain",
            "Location & Workplace Mode",
            "Date Applied",
            "Application Status",
            "ATS Match Score",
            "Key Matched Skills",
            "CV Template Used",
            "LinkedIn Job Posting URL"
        ]

        # Styling definitions
        header_fill = PatternFill(start_color="182B49", end_color="182B49", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=10, color="1F2937")
        alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        # Write Header Row
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        # Write Data Rows
        for row_idx, app in enumerate(applications, start=2):
            company = sanitize_excel_cell(app.get("company", "Company"))
            title = sanitize_excel_cell(app.get("job_title", "Position"))
            location = sanitize_excel_cell(app.get("location", "Worldwide Remote"))
            timestamp = sanitize_excel_cell(app.get("timestamp", datetime.now().strftime("%Y-%m-%d")))
            status = sanitize_excel_cell(app.get("status", "Applied").title())
            match_score = f"{app.get('ats_match_score', 95.0)}%"
            template = sanitize_excel_cell(app.get("template_used", "modern").replace("_", " ").title())
            job_url = sanitize_excel_cell(app.get("job_url", ""))
            
            # Skills extraction
            prefilled = app.get("prefilled_answers", {})
            skills_str = sanitize_excel_cell(prefilled.get("Key Matching Skills", "Python, Machine Learning, AI Engineering"))
            
            # Company Overview
            overview = sanitize_excel_cell(cls.get_company_overview(company, title))

            row_values = [
                company,
                title,
                overview,
                location,
                timestamp,
                status,
                match_score,
                skills_str,
                template,
                job_url
            ]

            is_alt = (row_idx % 2 == 0)
            for col_idx, val in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in (2, 3, 8)))
                if is_alt:
                    cell.fill = alt_row_fill
                if col_idx in (5, 6, 7, 9):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row_idx].height = 22

        # Auto-adjust column widths
        col_widths = {
            1: 22,  # Company
            2: 28,  # Role
            3: 38,  # Overview
            4: 22,  # Location
            5: 18,  # Date
            6: 18,  # Status
            7: 16,  # ATS
            8: 32,  # Skills
            9: 18,  # Template
            10: 34  # URL
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        wb.save(out_file)
        return out_file

    @classmethod
    def export_to_csv(
        cls,
        applications: List[Dict[str, Any]],
        output_path: Optional[Path] = None
    ) -> Path:
        """
        Builds a standard CSV file with company intelligence.
        """
        out_file = output_path or (OUTPUT_DIR / "Company_Applications_Tracker.csv")
        out_file.parent.mkdir(parents=True, exist_ok=True)

        headers = [
            "Company Name",
            "Job Title / Role",
            "Company Overview & Domain",
            "Location & Workplace Mode",
            "Date Applied",
            "Application Status",
            "ATS Match Score",
            "Key Matched Skills",
            "CV Template Used",
            "LinkedIn Job Posting URL"
        ]

        with open(out_file, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for app in applications:
                company = sanitize_excel_cell(app.get("company", "Company"))
                title = sanitize_excel_cell(app.get("job_title", "Position"))
                location = sanitize_excel_cell(app.get("location", "Worldwide Remote"))
                timestamp = sanitize_excel_cell(app.get("timestamp", datetime.now().strftime("%Y-%m-%d")))
                status = sanitize_excel_cell(app.get("status", "Applied").title())
                match_score = f"{app.get('ats_match_score', 95.0)}%"
                template = sanitize_excel_cell(app.get("template_used", "modern").replace("_", " ").title())
                job_url = sanitize_excel_cell(app.get("job_url", ""))
                prefilled = app.get("prefilled_answers", {})
                skills_str = sanitize_excel_cell(prefilled.get("Key Matching Skills", "Python, Machine Learning, AI Engineering"))
                overview = sanitize_excel_cell(cls.get_company_overview(company, title))

                writer.writerow([
                    company,
                    title,
                    overview,
                    location,
                    timestamp,
                    status,
                    match_score,
                    skills_str,
                    template,
                    job_url
                ])

        return out_file
